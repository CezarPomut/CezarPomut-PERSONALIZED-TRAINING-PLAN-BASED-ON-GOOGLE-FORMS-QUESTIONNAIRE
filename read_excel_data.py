import openpyxl
from select_training import *
from generate_training import *
import CTkMessagebox
from tkinter import messagebox


# path="C:/Users/pomut/Desktop/Sala.xlsx"


lista=[]
def extragere(string):
    numar = ""
    string=str(string)
    for x in range(0, len(string)):
        if x == 0 and string[x].isdecimal():
            numar = numar + string[x]
        if (x >= 1 and x < len(string) - 1) and string[x].isdecimal():
            numar = numar + string[x]
        if (x >= 1 and x < len(string) - 1) and (string[x] == "." or string[x] == ",") and string[x - 1].isdecimal() and \
                string[x + 1].isdecimal():
            if string[x] == ",":
                numar = numar + "."
            if string[x] == ".":
                numar = numar + string[x]
        if x == len(string) - 1 and string[x].isdecimal():
            numar = numar + string[x]
            return numar
        if (string[x].isalpha() or string[x].isspace()) and (string[x] != "," and string[x] != "."):
            if numar != "":
                return numar
    return numar

class Intrebari:
    def __init__(self,nume="", sex="", kg=0, inaltime=0.0, scop="",avansat="", perioada="",numar_sedinte=0, sanatate="", problema="",tipantrenament=""):
        self.__nume=nume
        self.__sex=sex
        self.__kg= kg
        self.__inaltime= inaltime
        self.__scop = scop
        self.__avansat =avansat
        self.__perioada = perioada
        self.__numar_sedinte = numar_sedinte
        self.__sanatate = sanatate
        self.__problema = problema
        self.__tip_antrenament= tipantrenament
        self.antrenament=[]

    def __repr__(self):
        return "("+self.__nume+","+self.__sex+","+ str(self.__kg)+","+str(self.__inaltime)+","+self.__scop+","+self.__avansat+","+self.__perioada+","+str(self.__numar_sedinte)+","+self.__sanatate+","+self.__problema+","+self.__tip_antrenament+")"

    def get_nume(self):
        return self.__nume

    def get_sex(self):
        return self.__sex

    def get_kg(self):
        return self.__kg

    def get_inaltime(self):
        return self.__inaltime

    def get_scop(self):
        return self.__scop

    def get_avansat(self):
        return self.__avansat

    def get_perioada(self):
        return self.__perioada

    def get_numar_sedinte(self):
        return self.__numar_sedinte

    def get_sanatate(self):
        return self.__sanatate

    def get_problema(self):
        return self.__problema

    def get_tip_antrenament(self):
        return self.__tip_antrenament

    def set_nume(self, x):
        self.__nume = x

    def set_sex(self, x):
        self.__sex = x

    def set_kg(self, x):
        self.__kg = x

    def set_inaltime(self, x):
        self.__inaltime = x

    def set_scop(self, x):
        self.__scop = x

    def set_avansat(self, x):
        self.__avansat = x

    def set_perioada(self, x):
        self.__perioada = x

    def set_numar_sedinte(self, x):
        self.__nr_sedinte = x

    def set_sanatate(self, x):
        self.__sanatate = x

    def set_problema(self, x):
        self.__problema = x

    def set_tip_antrenament(self, x):
        self.__tip_antrenament = x

    @classmethod
    def modificare(cls, lista_secundara):
        for x in lista_secundara:
            x.set_inaltime(extragere(x.get_inaltime()))
            x.set_kg(extragere(x.get_kg()))
            x.set_numar_sedinte(extragere(x.get_numar_sedinte()))

    @classmethod
    def read_excel(cls,file_path):
        try:
            wb_obj = openpyxl.load_workbook(file_path.strip())
        except:
            messagebox.showerror('Python Error', 'Error: Please check the path!')
        try:
            sheet1 = wb_obj["Form Responses 1"]
        except:
            messagebox.showerror('Python Error','Error: Please check if the file doesnt have Form Respomses 1')

        maxr = sheet1.max_row
        maxc = sheet1.max_column
        lista=[]
        for r in range(1, maxr + 1):
            nume=""
            sex=""
            kg=0
            inaltime=0.0
            scop=""
            avansat=""
            perioada=""
            numar_sedinte=0
            sanatate=""
            problema=""
            tip_antrenament=""
            if r!=1:
              for c in range(1, maxc + 1):
               if c!=1:
                 if c==2:
                   nume=sheet1.cell(row=r, column=c).value
                   #print(nume)
                 if c==3:
                   sex=sheet1.cell(row=r, column=c).value
                   #print(sex)
                 if c==4:
                   kg=sheet1.cell(row=r, column=c).value
                   #print(kg)
                 if c==5:
                   inaltime=sheet1.cell(row=r, column=c).value
                   #print(inaltime)
                 if c==6:
                    scop=sheet1.cell(row=r, column=c).value
                    #print(scop)
                 if c==7:
                     avansat=sheet1.cell(row=r, column=c).value
                     #print(avansat)
                 if c==8:
                     perioada=sheet1.cell(row=r, column=c).value
                     #print(perioada)
                 if c==9:
                     numar_sedinte=sheet1.cell(row=r, column=c).value
                     #print(numar_sedinte)
                 if c==10:
                     sanatate=sheet1.cell(row=r, column=c).value
                     #print(sanatate)
                 if c==11:
                     problema=sheet1.cell(row=r, column=c).value
                     if problema==None:
                         problema="nu are"
                     # print(problema)
                 if c==12:
                     tip_antrenament=sheet1.cell(row=r, column=c).value
                     # print(tip_antrenament)


              obiect=Intrebari(nume, sex, kg, inaltime, scop,avansat, perioada,numar_sedinte, sanatate, problema,tip_antrenament)
              #print(obiect)
              lista.append(obiect)
        wb_obj.save("Sala.xlsx")
        return lista


def run_main(file_path):
    lista_secundara=Intrebari.read_excel(file_path)
    # print(lista_secundara)

    # for x in range(0,len(lista_secundara)):
    #     print(lista_secundara[x].nume)

    for object in lista_secundara:
        print(object)

    Intrebari.modificare(lista_secundara)

    for x in lista_secundara:
        selectare_antrenament(x)

    for object in lista_secundara:
        print(object.antrenament)

    # print(lista_secundara[0].nume)
    for object in lista_secundara:
         generare_antrenament(object)


